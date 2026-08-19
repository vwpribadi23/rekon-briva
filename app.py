import streamlit as st
import pandas as pd
import re
import io
from openpyxl import Workbook


# ============================================================
# 1. PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Rekonsiliasi Bank Fastpay",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Rekonsiliasi Bank Fastpay")
st.write(
    "Dashboard rekonsiliasi otomatis antara data deposit FMSS "
    "dengan mutasi bank."
)

st.divider()


# ============================================================
# 2. SESSION STATE
# ============================================================

DEFAULT_STATE = {
    "sudah_diproses": False,
    "df_matched": pd.DataFrame(),
    "df_issue_fmss": pd.DataFrame(),
    "df_issue_bank": pd.DataFrame(),
    "df_invalid_fmss": pd.DataFrame(),
    "df_invalid_bank": pd.DataFrame(),
    "pilihan_bank_terakhir": "",
}

for key, value in DEFAULT_STATE.items():
    if key not in st.session_state:
        st.session_state[key] = value


# ============================================================
# 3. HELPER FUNCTIONS
# ============================================================

def reset_result():
    st.session_state.sudah_diproses = False
    st.session_state.df_matched = pd.DataFrame()
    st.session_state.df_issue_fmss = pd.DataFrame()
    st.session_state.df_issue_bank = pd.DataFrame()
    st.session_state.df_invalid_fmss = pd.DataFrame()
    st.session_state.df_invalid_bank = pd.DataFrame()


def read_uploaded_file(uploaded_file):
    """
    Membaca CSV/XLSX secara fleksibel.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    uploaded_file.seek(0)

    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        try:
            return pd.read_csv(
                uploaded_file,
                sep=None,
                engine="python",
                dtype=str
            )
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(
                uploaded_file,
                dtype=str
            )

    elif filename.endswith(".xlsx"):
        return pd.read_excel(
            uploaded_file,
            dtype=str
        )

    raise ValueError(
        f"Format file tidak didukung: {uploaded_file.name}"
    )


def normalize_columns(df):
    """
    Normalisasi nama kolom agar lebih aman.
    """
    df = df.copy()

    df.columns = [
        str(col).strip()
        for col in df.columns
    ]

    return df


def find_column(df, candidates):
    """
    Mencari nama kolom berdasarkan kandidat.
    """
    normalized = {
        str(col).strip().lower(): col
        for col in df.columns
    }

    for candidate in candidates:
        key = candidate.strip().lower()

        if key in normalized:
            return normalized[key]

    return None


def numeric_value(series):
    """
    Konversi angka dengan aman.
    Mendukung format:
    1,000
    1.000
    1000
    Rp 1.000
    """
    if series is None:
        return pd.Series(dtype=float)

    s = series.astype(str).str.strip()

    s = (
        s.str.replace("Rp", "", regex=False)
         .str.replace("rp", "", regex=False)
         .str.replace(" ", "", regex=False)
    )

    # Jika terdapat format Indonesia 1.999.000
    # ubah menjadi 1999000
    s = s.str.replace(",", "", regex=False)
    s = s.str.replace(".", "", regex=False)

    return pd.to_numeric(
        s,
        errors="coerce"
    ).fillna(0)


def extract_va(text):
    """
    Mengenali dua prefix BRIVA:

    57888 = BRIVA FASTPAY
    57708 = BRIVA RAJABILLER
    """

    if pd.isna(text):
        return None

    text = str(text)

    match = re.search(
        r"(57888\d{5,15}|57708\d{5,15})",
        text
    )

    if match:
        return match.group(1)

    return None


def classify_va(va):
    """
    Menentukan jenis VA.
    """

    if pd.isna(va) or va is None:
        return "TIDAK TERIDENTIFIKASI"

    va = str(va)

    if va.startswith("57888"):
        return "BRIVA FASTPAY"

    if va.startswith("57708"):
        return "BRIVA RAJABILLER"

    return "TIDAK TERIDENTIFIKASI"


def extract_va_by_prefix(text, prefix):
    """
    Ekstraksi VA untuk file bank tertentu.
    """

    if pd.isna(text):
        return None

    text = str(text)

    match = re.search(
        rf"({prefix}\d{{5,15}})",
        text
    )

    if match:
        return match.group(1)

    return None


def safe_prepare_excel_df(df):
    """
    Membersihkan dataframe sebelum dimasukkan ke Excel.

    Ini penting untuk mencegah error ExcelWriter/openpyxl,
    terutama jika terdapat datetime, NaN, Inf, atau timezone.
    """

    if df is None or df.empty:
        return pd.DataFrame({"INFO": ["Tidak ada data"]})

    result = df.copy()

    for col in result.columns:

        # Datetime
        if pd.api.types.is_datetime64_any_dtype(result[col]):
            try:
                result[col] = result[col].dt.tz_localize(None)
            except Exception:
                pass

            result[col] = result[col].astype(str)

        # Object berisi timezone datetime
        if result[col].dtype == "object":
            try:
                result[col] = result[col].apply(
                    lambda x: (
                        x.tz_localize(None)
                        if hasattr(x, "tzinfo")
                        and x.tzinfo is not None
                        else x
                    )
                )
            except Exception:
                pass

        # Inf
        result[col] = result[col].replace(
            [float("inf"), float("-inf")],
            None
        )

    return result


def dataframe_to_excel(dataframes):
    """
    Membuat file Excel dengan openpyxl secara aman.

    dataframes:
    {
        "MATCHED_OK": df,
        "ISSUE_FMSS": df,
        ...
    }
    """

    output = io.BytesIO()

    wb = Workbook()

    # Hapus sheet default
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, df in dataframes.items():

        # Excel maksimal 31 karakter untuk nama sheet
        sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        clean_df = safe_prepare_excel_df(df)

        # Header
        for col_idx, column in enumerate(
            clean_df.columns,
            start=1
        ):
            ws.cell(
                row=1,
                column=col_idx,
                value=str(column)
            )

        # Data
        for row_idx, row in enumerate(
            clean_df.itertuples(index=False, name=None),
            start=2
        ):
            for col_idx, value in enumerate(
                row,
                start=1
            ):

                if pd.isna(value):
                    value = None

                # Hindari object aneh masuk Excel
                if not isinstance(
                    value,
                    (
                        str,
                        int,
                        float,
                        bool,
                        type(None)
                    )
                ):
                    value = str(value)

                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value
                )

        # Auto width
        for column_cells in ws.columns:

            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:

                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 10),
                50
            )

    wb.save(output)

    output.seek(0)

    return output.getvalue()


def make_issue_table(df, issue_type):
    """
    Membuat tabel ringkas untuk tampilan dashboard.
    """

    if df is None or df.empty:
        return pd.DataFrame()

    result = pd.DataFrame()

    if "VA" in df.columns:
        result["KODE VA"] = df["VA"]

    if "JENIS VA" in df.columns:
        result["JENIS VA"] = df["JENIS VA"]

    if "nominal" in df.columns:
        result["NOMINAL"] = df["nominal"]

    elif "MUTASI_KREDIT_NUM" in df.columns:
        result["NOMINAL"] = df["MUTASI_KREDIT_NUM"]

    if "ISSUE" in df.columns:
        result["ISSUE"] = df["ISSUE"]

    else:
        result["ISSUE"] = issue_type

    return result


# ============================================================
# 4. PENGATURAN BANK
# ============================================================

st.subheader("1. Pengaturan Data")

opsi_bank = [
    "",
    "BRIVA",
    "BNIVA",
    "BCAVA",
    "MANDIRIVA",
    "BSIVA",
    "MuamalatVA"
]

pilihan_bank = st.selectbox(
    "Pilih Bank Sumber Mutasi:",
    opsi_bank
)


# Reset jika bank berubah
if (
    pilihan_bank
    != st.session_state.pilihan_bank_terakhir
):

    reset_result()

    st.session_state.pilihan_bank_terakhir = (
        pilihan_bank
    )


# ============================================================
# 5. UPLOAD FILE
# ============================================================

st.subheader("2. Unggah File")


if pilihan_bank == "BRIVA":

    # ========================================================
    # BRIVA → 3 FILE
    # ========================================================

    col1, col2, col3 = st.columns(3)

    with col1:

        st.markdown("### 📄 FMSS")

        file_int = st.file_uploader(
            "Upload data FMSS",
            type=["csv", "xlsx"],
            key="fmss_briva"
        )

    with col2:

        st.markdown("### 🏦 BRIVA Fastpay")

        st.caption("Prefix VA: 57888")

        file_bnk_57888 = st.file_uploader(
            "Upload mutasi BRIVA 57888",
            type=["csv", "xlsx"],
            key="briva_57888"
        )

    with col3:

        st.markdown("### 🏦 BRIVA Rajabiller")

        st.caption("Prefix VA: 57708")

        file_bnk_57708 = st.file_uploader(
            "Upload mutasi BRIVA 57708",
            type=["csv", "xlsx"],
            key="briva_57708"
        )

    # ========================================================
    # FEE
    # ========================================================

    st.subheader("3. Konfigurasi Fee")

    fee_col1, fee_col2 = st.columns(2)

    with fee_col1:

        fee_57888 = st.number_input(
            "Fee BRIVA Fastpay (57888)",
            min_value=0,
            value=1000,
            step=100
        )

    with fee_col2:

        fee_57708 = st.number_input(
            "Fee BRIVA Rajabiller (57708)",
            min_value=0,
            value=1000,
            step=100
        )

    st.caption(
        "Rumus matching: Nominal FMSS + Fee = Nominal Mutasi Bank"
    )

    semua_file_ada = (
        file_int
        and file_bnk_57888
        and file_bnk_57708
    )


else:

    # ========================================================
    # BANK LAIN → 2 FILE
    # ========================================================

    col1, col2 = st.columns(2)

    with col1:

        file_int = st.file_uploader(
            "Upload data FMSS",
            type=["csv", "xlsx"],
            key="fmss_other"
        )

    with col2:

        label_bank = (
            f"Upload Mutasi Bank ({pilihan_bank})"
            if pilihan_bank
            else
            "Upload Mutasi Bank"
        )

        file_bnk = st.file_uploader(
            label_bank,
            type=["csv", "xlsx"],
            key="bank_other"
        )

    semua_file_ada = (
        pilihan_bank != ""
        and file_int
        and file_bnk
    )


# ============================================================
# 6. PROSES REKONSILIASI BRIVA
# ============================================================

if pilihan_bank == "BRIVA" and semua_file_ada:

    st.divider()

    if st.button(
        "🚀 Mulai Croscek Data BRIVA",
        type="primary",
        use_container_width=True
    ):

        reset_result()

        with st.spinner(
            "Sedang memproses rekonsiliasi BRIVA..."
        ):

            try:

                # ====================================================
                # READ FILE
                # ====================================================

                df_int = normalize_columns(
                    read_uploaded_file(file_int)
                )

                df_bank_57888 = normalize_columns(
                    read_uploaded_file(file_bnk_57888)
                )

                df_bank_57708 = normalize_columns(
                    read_uploaded_file(file_bnk_57708)
                )


                # ====================================================
                # VALIDASI KOLOM FMSS
                # ====================================================

                status_col = find_column(
                    df_int,
                    [
                        "status"
                    ]
                )

                keterangan_col = find_column(
                    df_int,
                    [
                        "keterangan",
                        "deskripsi",
                        "description"
                    ]
                )

                nominal_col = find_column(
                    df_int,
                    [
                        "nominal",
                        "amount"
                    ]
                )

                if not status_col:
                    raise ValueError(
                        "Kolom 'status' tidak ditemukan di file FMSS."
                    )

                if not keterangan_col:
                    raise ValueError(
                        "Kolom 'keterangan' tidak ditemukan di file FMSS."
                    )

                if not nominal_col:
                    raise ValueError(
                        "Kolom 'nominal' tidak ditemukan di file FMSS."
                    )


                # ====================================================
                # FMSS SUCCESS
                # ====================================================

                df_int_success = df_int[
                    df_int[status_col]
                    .astype(str)
                    .str.upper()
                    .str.strip()
                    == "SUKSES"
                ].copy()


                # ====================================================
                # EXTRACT VA FMSS
                # ====================================================

                df_int_success["VA"] = (
                    df_int_success[keterangan_col]
                    .apply(extract_va)
                )

                df_int_success["JENIS VA"] = (
                    df_int_success["VA"]
                    .apply(classify_va)
                )

                df_int_success["nominal"] = (
                    numeric_value(
                        df_int_success[nominal_col]
                    )
                )


                # ====================================================
                # FMSS INVALID VA
                #
                # Transaksi SUKSES tetapi tidak menemukan
                # prefix 57888 / 57708
                # ====================================================

                df_invalid_fmss = (
                    df_int_success[
                        df_int_success["VA"].isna()
                    ]
                    .copy()
                )

                if not df_invalid_fmss.empty:

                    df_invalid_fmss["ISSUE"] = (
                        "VA_TIDAK_TERIDENTIFIKASI"
                    )


                # Hanya transaksi yang punya VA valid
                df_int_valid = (
                    df_int_success[
                        df_int_success["VA"].notna()
                    ]
                    .copy()
                )


                # ====================================================
                # PROCESS BANK 57888
                # ====================================================

                desk_57888 = find_column(
                    df_bank_57888,
                    [
                        "DESK_TRAN",
                        "desk_tran",
                        "keterangan",
                        "deskripsi"
                    ]
                )

                credit_57888 = find_column(
                    df_bank_57888,
                    [
                        "MUTASI_KREDIT",
                        "mutasi_kredit",
                        "kredit",
                        "credit"
                    ]
                )

                if not desk_57888:
                    raise ValueError(
                        "Kolom DESK_TRAN/keterangan tidak ditemukan "
                        "pada file mutasi BRIVA 57888."
                    )

                if not credit_57888:
                    raise ValueError(
                        "Kolom MUTASI_KREDIT/kredit tidak ditemukan "
                        "pada file mutasi BRIVA 57888."
                    )

                df_bank_57888["MUTASI_KREDIT_NUM"] = (
                    numeric_value(
                        df_bank_57888[credit_57888]
                    )
                )

                df_bank_57888 = df_bank_57888[
                    df_bank_57888["MUTASI_KREDIT_NUM"] > 0
                ].copy()

                df_bank_57888["VA"] = (
                    df_bank_57888[desk_57888]
                    .apply(
                        lambda x:
                        extract_va_by_prefix(
                            x,
                            "57888"
                        )
                    )
                )

                df_bank_57888["JENIS VA"] = (
                    "BRIVA FASTPAY"
                )


                # ====================================================
                # BANK INVALID 57888
                # ====================================================

                invalid_bank_57888 = (
                    df_bank_57888[
                        df_bank_57888["VA"].isna()
                    ]
                    .copy()
                )

                if not invalid_bank_57888.empty:

                    invalid_bank_57888["ISSUE"] = (
                        "VA_TIDAK_TERIDENTIFIKASI"
                    )


                df_bank_57888_valid = (
                    df_bank_57888[
                        df_bank_57888["VA"].notna()
                    ]
                    .copy()
                )


                # ====================================================
                # PROCESS BANK 57708
                # ====================================================

                desk_57708 = find_column(
                    df_bank_57708,
                    [
                        "DESK_TRAN",
                        "desk_tran",
                        "keterangan",
                        "deskripsi"
                    ]
                )

                credit_57708 = find_column(
                    df_bank_57708,
                    [
                        "MUTASI_KREDIT",
                        "mutasi_kredit",
                        "kredit",
                        "credit"
                    ]
                )

                if not desk_57708:
                    raise ValueError(
                        "Kolom DESK_TRAN/keterangan tidak ditemukan "
                        "pada file mutasi BRIVA 57708."
                    )

                if not credit_57708:
                    raise ValueError(
                        "Kolom MUTASI_KREDIT/kredit tidak ditemukan "
                        "pada file mutasi BRIVA 57708."
                    )

                df_bank_57708["MUTASI_KREDIT_NUM"] = (
                    numeric_value(
                        df_bank_57708[credit_57708]
                    )
                )

                df_bank_57708 = df_bank_57708[
                    df_bank_57708["MUTASI_KREDIT_NUM"] > 0
                ].copy()

                df_bank_57708["VA"] = (
                    df_bank_57708[desk_57708]
                    .apply(
                        lambda x:
                        extract_va_by_prefix(
                            x,
                            "57708"
                        )
                    )
                )

                df_bank_57708["JENIS VA"] = (
                    "BRIVA RAJABILLER"
                )


                # ====================================================
                # BANK INVALID 57708
                # ====================================================

                invalid_bank_57708 = (
                    df_bank_57708[
                        df_bank_57708["VA"].isna()
                    ]
                    .copy()
                )

                if not invalid_bank_57708.empty:

                    invalid_bank_57708["ISSUE"] = (
                        "VA_TIDAK_TERIDENTIFIKASI"
                    )


                df_bank_57708_valid = (
                    df_bank_57708[
                        df_bank_57708["VA"].notna()
                    ]
                    .copy()
                )


                # ====================================================
                # MATCHING FUNCTION
                # ====================================================

                def reconcile(
                    df_fmss,
                    df_bank,
                    fee,
                    jenis_va
                ):

                    fmss = df_fmss[
                        df_fmss["JENIS VA"]
                        == jenis_va
                    ].copy()

                    bank = df_bank.copy()

                    if fmss.empty:

                        return (
                            pd.DataFrame(),
                            pd.DataFrame(),
                            bank
                        )

                    if bank.empty:

                        fmss["ISSUE"] = "FMSS_ONLY"

                        return (
                            pd.DataFrame(),
                            fmss,
                            pd.DataFrame()
                        )


                    # -----------------------------------------------
                    # NOMINAL ADJUSTMENT
                    # -----------------------------------------------

                    fmss["NOMINAL_INT_ADJ"] = (
                        fmss["nominal"] + fee
                    )


                    # -----------------------------------------------
                    # OCCURRENCE NUMBER
                    #
                    # Digunakan agar transaksi duplicate tetap
                    # dicocokkan 1-to-1.
                    # -----------------------------------------------

                    fmss["_MATCH_NO"] = (
                        fmss
                        .groupby(
                            [
                                "VA",
                                "NOMINAL_INT_ADJ"
                            ],
                            dropna=False
                        )
                        .cumcount()
                    )

                    bank["_MATCH_NO"] = (
                        bank
                        .groupby(
                            [
                                "VA",
                                "MUTASI_KREDIT_NUM"
                            ],
                            dropna=False
                        )
                        .cumcount()
                    )


                    # -----------------------------------------------
                    # MERGE 1-to-1
                    # -----------------------------------------------

                    matched = pd.merge(
                        fmss,
                        bank,
                        left_on=[
                            "VA",
                            "NOMINAL_INT_ADJ",
                            "_MATCH_NO"
                        ],
                        right_on=[
                            "VA",
                            "MUTASI_KREDIT_NUM",
                            "_MATCH_NO"
                        ],
                        how="inner",
                        suffixes=(
                            "_FMSS",
                            "_BANK"
                        )
                    )


                    # -----------------------------------------------
                    # FMSS UNMATCHED
                    # -----------------------------------------------

                    fmss_key = set(
                        zip(
                            fmss["VA"],
                            fmss["NOMINAL_INT_ADJ"],
                            fmss["_MATCH_NO"]
                        )
                    )

                    matched_key = set(
                        zip(
                            matched["VA"],
                            matched["NOMINAL_INT_ADJ"],
                            matched["_MATCH_NO"]
                        )
                    )

                    unmatched_fmss_key = (
                        fmss_key - matched_key
                    )

                    unmatched_fmss = fmss[
                        fmss.apply(
                            lambda row:
                            (
                                row["VA"],
                                row["NOMINAL_INT_ADJ"],
                                row["_MATCH_NO"]
                            )
                            in unmatched_fmss_key,
                            axis=1
                        )
                    ].copy()


                    # -----------------------------------------------
                    # BANK UNMATCHED
                    # -----------------------------------------------

                    bank_key = set(
                        zip(
                            bank["VA"],
                            bank["MUTASI_KREDIT_NUM"],
                            bank["_MATCH_NO"]
                        )
                    )

                    matched_bank_key = set(
                        zip(
                            matched["VA"],
                            matched["MUTASI_KREDIT_NUM"],
                            matched["_MATCH_NO"]
                        )
                    )

                    unmatched_bank_key = (
                        bank_key - matched_bank_key
                    )

                    unmatched_bank = bank[
                        bank.apply(
                            lambda row:
                            (
                                row["VA"],
                                row["MUTASI_KREDIT_NUM"],
                                row["_MATCH_NO"]
                            )
                            in unmatched_bank_key,
                            axis=1
                        )
                    ].copy()


                    # -----------------------------------------------
                    # ISSUE
                    # -----------------------------------------------

                    if not unmatched_fmss.empty:
                        unmatched_fmss["ISSUE"] = (
                            "FMSS_ONLY"
                        )

                    if not unmatched_bank.empty:
                        unmatched_bank["ISSUE"] = (
                            "BANK_ONLY"
                        )


                    # -----------------------------------------------
                    # MATCHED
                    # -----------------------------------------------

                    if not matched.empty:

                        matched["JENIS VA"] = (
                            jenis_va
                        )

                        matched["MATCH_STATUS"] = (
                            "MATCHED_OK"
                        )

                        matched["FEE"] = fee


                    # Cleanup
                    drop_cols = [
                        "_MATCH_NO"
                    ]

                    matched = matched.drop(
                        columns=drop_cols,
                        errors="ignore"
                    )

                    unmatched_fmss = unmatched_fmss.drop(
                        columns=drop_cols,
                        errors="ignore"
                    )

                    unmatched_bank = unmatched_bank.drop(
                        columns=drop_cols,
                        errors="ignore"
                    )

                    return (
                        matched,
                        unmatched_fmss,
                        unmatched_bank
                    )


                # ====================================================
                # RECONCILE FASTPAY
                # ====================================================

                (
                    matched_57888,
                    issue_fmss_57888,
                    issue_bank_57888
                ) = reconcile(
                    df_int_valid,
                    df_bank_57888_valid,
                    fee_57888,
                    "BRIVA FASTPAY"
                )


                # ====================================================
                # RECONCILE RAJABILLER
                # ====================================================

                (
                    matched_57708,
                    issue_fmss_57708,
                    issue_bank_57708
                ) = reconcile(
                    df_int_valid,
                    df_bank_57708_valid,
                    fee_57708,
                    "BRIVA RAJABILLER"
                )


                # ====================================================
                # COMBINE RESULT
                # ====================================================

                matched_all = pd.concat(
                    [
                        matched_57888,
                        matched_57708
                    ],
                    ignore_index=True
                )

                issue_fmss_all = pd.concat(
                    [
                        issue_fmss_57888,
                        issue_fmss_57708
                    ],
                    ignore_index=True
                )

                issue_bank_all = pd.concat(
                    [
                        issue_bank_57888,
                        issue_bank_57708
                    ],
                    ignore_index=True
                )

                invalid_fmss_all = (
                    df_invalid_fmss
                    .copy()
                )

                invalid_bank_all = pd.concat(
                    [
                        invalid_bank_57888,
                        invalid_bank_57708
                    ],
                    ignore_index=True
                )


                # ====================================================
                # SAVE SESSION STATE
                # ====================================================

                st.session_state.df_matched = (
                    matched_all
                )

                st.session_state.df_issue_fmss = (
                    issue_fmss_all
                )

                st.session_state.df_issue_bank = (
                    issue_bank_all
                )

                st.session_state.df_invalid_fmss = (
                    invalid_fmss_all
                )

                st.session_state.df_invalid_bank = (
                    invalid_bank_all
                )

                st.session_state.sudah_diproses = True


                st.success(
                    "Rekonsiliasi BRIVA selesai."
                )


            except Exception as e:

                st.session_state.sudah_diproses = False

                st.error(
                    "Terjadi kesalahan saat memproses data."
                )

                st.exception(e)


# ============================================================
# 7. BANK LAIN
# ============================================================

elif (
    pilihan_bank != ""
    and pilihan_bank != "BRIVA"
    and semua_file_ada
):

    st.divider()

    if st.button(
        f"🚀 Mulai Croscek Data {pilihan_bank}",
        type="primary",
        use_container_width=True
    ):

        st.warning(
            f"Modul rekonsiliasi {pilihan_bank} "
            "belum diaktifkan. Struktur upload sudah "
            "disiapkan agar bank ini dapat ditambahkan "
            "tanpa mengubah dashboard utama."
        )


# ============================================================
# 8. HASIL REKONSILIASI BRIVA
# ============================================================

if (
    pilihan_bank == "BRIVA"
    and st.session_state.sudah_diproses
):

    df_matched = (
        st.session_state.df_matched
    )

    df_issue_fmss = (
        st.session_state.df_issue_fmss
    )

    df_issue_bank = (
        st.session_state.df_issue_bank
    )

    df_invalid_fmss = (
        st.session_state.df_invalid_fmss
    )

    df_invalid_bank = (
        st.session_state.df_invalid_bank
    )


    # ========================================================
    # SUMMARY
    # ========================================================

    st.divider()

    st.subheader(
        "🎯 Ringkasan Rekonsiliasi BRIVA"
    )

    m1, m2, m3, m4 = st.columns(4)

    m1.metric(
        "✅ Matched Sempurna",
        f"{len(df_matched):,} Trx"
    )

    m2.metric(
        "⚠️ Issue FMSS",
        f"{len(df_issue_fmss):,} Trx"
    )

    m3.metric(
        "⚠️ Issue Bank",
        f"{len(df_issue_bank):,} Trx"
    )

    m4.metric(
        "🚨 Invalid VA",
        f"{len(df_invalid_fmss) + len(df_invalid_bank):,} Trx"
    )


    # ========================================================
    # AMOUNT SUMMARY
    # ========================================================

    st.subheader(
        "💰 Ringkasan Nominal"
    )

    amount_col1, amount_col2, amount_col3 = (
        st.columns(3)
    )


    matched_amount = 0

    if (
        not df_matched.empty
        and "nominal_FMSS" in df_matched.columns
    ):
        matched_amount = (
            pd.to_numeric(
                df_matched["nominal_FMSS"],
                errors="coerce"
            )
            .fillna(0)
            .sum()
        )

    issue_fmss_amount = 0

    if (
        not df_issue_fmss.empty
        and "nominal" in df_issue_fmss.columns
    ):
        issue_fmss_amount = (
            pd.to_numeric(
                df_issue_fmss["nominal"],
                errors="coerce"
            )
            .fillna(0)
            .sum()
        )

    issue_bank_amount = 0

    if (
        not df_issue_bank.empty
        and "MUTASI_KREDIT_NUM" in df_issue_bank.columns
    ):
        issue_bank_amount = (
            pd.to_numeric(
                df_issue_bank["MUTASI_KREDIT_NUM"],
                errors="coerce"
            )
            .fillna(0)
            .sum()
        )


    amount_col1.metric(
        "Matched",
        f"Rp {matched_amount:,.0f}"
    )

    amount_col2.metric(
        "Issue FMSS",
        f"Rp {issue_fmss_amount:,.0f}"
    )

    amount_col3.metric(
        "Issue Bank",
        f"Rp {issue_bank_amount:,.0f}"
    )


    # ========================================================
    # ISSUE TABLE
    # ========================================================

    st.divider()

    col_left, col_right = st.columns(2)


    # ========================================================
    # ISSUE FMSS
    # ========================================================

    with col_left:

        st.subheader(
            "🚨 Issue FMSS"
        )

        if not df_issue_fmss.empty:

            tampil_fmss = make_issue_table(
                df_issue_fmss,
                "FMSS_ONLY"
            )

            st.dataframe(
                tampil_fmss,
                use_container_width=True,
                hide_index=True
            )

        else:

            st.success(
                "Tidak ada issue FMSS. "
                "Semua transaksi FMSS memiliki pasangan."
            )


    # ========================================================
    # ISSUE BANK
    # ========================================================

    with col_right:

        st.subheader(
            "🚨 Issue Bank"
        )

        if not df_issue_bank.empty:

            tampil_bank = make_issue_table(
                df_issue_bank,
                "BANK_ONLY"
            )

            st.dataframe(
                tampil_bank,
                use_container_width=True,
                hide_index=True
            )

        else:

            st.success(
                "Tidak ada issue Bank. "
                "Semua transaksi bank memiliki pasangan."
            )


    # ========================================================
    # INVALID VA
    # ========================================================

    st.divider()

    with st.expander(
        "⚠️ Transaksi dengan VA Tidak Teridentifikasi",
        expanded=False
    ):

        invalid_col1, invalid_col2 = (
            st.columns(2)
        )


        # ----------------------------------------------------
        # INVALID FMSS
        # ----------------------------------------------------

        with invalid_col1:

            st.subheader(
                "FMSS Invalid VA"
            )

            if not df_invalid_fmss.empty:

                invalid_fmss_display = (
                    df_invalid_fmss.copy()
                )

                columns_to_show = []

                for col in [
                    "tanggal_transfer",
                    "keterangan",
                    "nominal",
                    "ISSUE"
                ]:

                    if col in invalid_fmss_display.columns:
                        columns_to_show.append(col)

                if columns_to_show:

                    st.dataframe(
                        invalid_fmss_display[
                            columns_to_show
                        ],
                        use_container_width=True,
                        hide_index=True
                    )

                else:

                    st.dataframe(
                        invalid_fmss_display,
                        use_container_width=True,
                        hide_index=True
                    )

            else:

                st.success(
                    "Tidak ada FMSS invalid VA."
                )


        # ----------------------------------------------------
        # INVALID BANK
        # ----------------------------------------------------

        with invalid_col2:

            st.subheader(
                "Bank Invalid VA"
            )

            if not df_invalid_bank.empty:

                invalid_bank_display = (
                    df_invalid_bank.copy()
                )

                columns_to_show = []

                for col in [
                    "DESK_TRAN",
                    "keterangan",
                    "MUTASI_KREDIT_NUM",
                    "ISSUE"
                ]:

                    if col in invalid_bank_display.columns:
                        columns_to_show.append(col)

                if columns_to_show:

                    st.dataframe(
                        invalid_bank_display[
                            columns_to_show
                        ],
                        use_container_width=True,
                        hide_index=True
                    )

                else:

                    st.dataframe(
                        invalid_bank_display,
                        use_container_width=True,
                        hide_index=True
                    )

            else:

                st.success(
                    "Tidak ada Bank invalid VA."
                )


    # ========================================================
    # BREAKDOWN JENIS VA
    # ========================================================

    st.divider()

    st.subheader(
        "📊 Breakdown Rekonsiliasi"
    )

    breakdown_rows = []

    for jenis in [
        "BRIVA FASTPAY",
        "BRIVA RAJABILLER"
    ]:

        matched_count = 0
        fmss_count = 0
        bank_count = 0

        matched_nominal = 0
        fmss_nominal = 0
        bank_nominal = 0


        if not df_matched.empty:

            temp = df_matched[
                df_matched["JENIS VA"]
                == jenis
            ]

            matched_count = len(temp)

            if "nominal_FMSS" in temp.columns:

                matched_nominal = (
                    pd.to_numeric(
                        temp["nominal_FMSS"],
                        errors="coerce"
                    )
                    .fillna(0)
                    .sum()
                )


        if not df_issue_fmss.empty:

            temp = df_issue_fmss[
                df_issue_fmss["JENIS VA"]
                == jenis
            ]

            fmss_count = len(temp)

            if "nominal" in temp.columns:

                fmss_nominal = (
                    pd.to_numeric(
                        temp["nominal"],
                        errors="coerce"
                    )
                    .fillna(0)
                    .sum()
                )


        if not df_issue_bank.empty:

            temp = df_issue_bank[
                df_issue_bank["JENIS VA"]
                == jenis
            ]

            bank_count = len(temp)

            if "MUTASI_KREDIT_NUM" in temp.columns:

                bank_nominal = (
                    pd.to_numeric(
                        temp["MUTASI_KREDIT_NUM"],
                        errors="coerce"
                    )
                    .fillna(0)
                    .sum()
                )


        breakdown_rows.append(
            {
                "JENIS VA": jenis,
                "MATCHED": matched_count,
                "ISSUE FMSS": fmss_count,
                "ISSUE BANK": bank_count,
                "NOMINAL MATCHED": matched_nominal,
                "NOMINAL ISSUE FMSS": fmss_nominal,
                "NOMINAL ISSUE BANK": bank_nominal,
            }
        )


    df_breakdown = pd.DataFrame(
        breakdown_rows
    )

    st.dataframe(
        df_breakdown,
        use_container_width=True,
        hide_index=True
    )


    # ========================================================
    # DOWNLOAD EXCEL
    # ========================================================

    st.divider()

    st.subheader(
        "📥 Download Laporan"
    )

    excel_data = dataframe_to_excel(
        {
            "MATCHED_OK": df_matched,
            "ISSUE_FMSS": df_issue_fmss,
            "ISSUE_BANK": df_issue_bank,
            "FMSS_INVALID_VA": df_invalid_fmss,
            "BANK_INVALID_VA": df_invalid_bank,
            "BREAKDOWN": df_breakdown,
        }
    )

    st.download_button(
        label="📥 Download Laporan Lengkap (.xlsx)",
        data=excel_data,
        file_name="Laporan_Rekonsiliasi_BRIVA.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        type="primary",
        use_container_width=True
    )


# ============================================================
# 9. INFO JIKA BELUM PILIH BANK
# ============================================================

elif (
    pilihan_bank == ""
    and (
        st.session_state.get(
            "pilihan_bank_terakhir",
            ""
        ) == ""
    )
):

    st.info(
        "💡 Silakan pilih **Bank Sumber Mutasi** "
        "terlebih dahulu pada dropdown di atas."
    )
